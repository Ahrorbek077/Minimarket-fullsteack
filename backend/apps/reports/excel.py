"""
Excel export — openpyxl bilan.
Har bir hisobot turi uchun alohida funksiya.
"""
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter


# ─── Style helpers ─────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(bold=True, size=14, color="1F4E79")
TOTAL_FILL    = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT    = Font(bold=True, size=11)
ALT_FILL      = PatternFill("solid", fgColor="EBF3FB")
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
RIGHT         = Alignment(horizontal="right",  vertical="center")
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _set_header_row(ws, row: int, headers: list[tuple]):
    """
    headers = [("Sarlavha", width), ...]
    """
    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width


def _set_title(ws, title: str, subtitle: str = ""):
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    cell = ws["A1"]
    cell.value     = title
    cell.font      = TITLE_FONT
    cell.alignment = LEFT
    if subtitle:
        sub = ws["A2"]
        sub.value     = subtitle
        sub.font      = Font(size=10, color="666666", italic=True)
        sub.alignment = LEFT


def _data_cell(ws, row, col, value, align=LEFT, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = align
    cell.border    = THIN_BORDER
    if number_format:
        cell.number_format = number_format
    # Alternating rows
    if row % 2 == 0:
        cell.fill = ALT_FILL
    return cell


def _total_row(ws, row: int, cols_values: list):
    for col, value in cols_values:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font      = TOTAL_FONT
        cell.fill      = TOTAL_FILL
        cell.alignment = RIGHT
        cell.border    = THIN_BORDER


def _to_buffer(wb) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Sales Excel ───────────────────────────────────────────────────────────────

def export_sales_excel(sales_qs, date_from, date_to, summary: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Sotuvlar ro'yxati ──────────────────────────────────────
    ws = wb.active
    ws.title = "Sotuvlar"
    _set_title(ws, "Sotuvlar hisoboti", f"{date_from} — {date_to}")

    headers = [
        ("№",            5),  ("Chek raqami",  18), ("Sana",          14),
        ("Kassir",       20), ("Jami summa",   14), ("Chegirma",      12),
        ("To'lanadigan", 14), ("Naqd",         12), ("Karta",         12),
        ("Nasiya",       12), ("Holat",        12),
    ]
    _set_header_row(ws, 4, headers)

    from sales.models import PaymentMethod
    row = 5
    for idx, sale in enumerate(sales_qs, 1):
        cash = sum(p.amount for p in sale.payments.all() if p.method == PaymentMethod.CASH)
        card = sum(p.amount for p in sale.payments.all() if p.method == PaymentMethod.CARD)
        debt = sum(p.amount for p in sale.payments.all() if p.method == PaymentMethod.DEBT)

        _data_cell(ws, row, 1,  idx,                                    CENTER)
        _data_cell(ws, row, 2,  sale.invoice_no,                        LEFT)
        _data_cell(ws, row, 3,  sale.created_at.strftime("%d.%m.%Y %H:%M"), CENTER)
        _data_cell(ws, row, 4,  sale.cashier.full_name if sale.cashier else "", LEFT)
        _data_cell(ws, row, 5,  float(sale.total_amount),               RIGHT, '#,##0.00')
        _data_cell(ws, row, 6,  float(sale.discount_amount),            RIGHT, '#,##0.00')
        _data_cell(ws, row, 7,  float(sale.net_amount),                 RIGHT, '#,##0.00')
        _data_cell(ws, row, 8,  float(cash),                            RIGHT, '#,##0.00')
        _data_cell(ws, row, 9,  float(card),                            RIGHT, '#,##0.00')
        _data_cell(ws, row, 10, float(debt),                            RIGHT, '#,##0.00')
        _data_cell(ws, row, 11, sale.get_status_display(),              CENTER)
        row += 1

    # Jami qator
    _total_row(ws, row, [
        (1,  "JAMI"),
        (4,  f"{summary['total_sales']} ta sotuv"),
        (5,  float(summary["total_revenue"])),
        (6,  float(summary["total_discount"])),
        (7,  float(summary["total_revenue"])),
        (8,  float(summary["cash_total"])),
        (9,  float(summary["card_total"])),
        (10, float(summary["debt_total"])),
    ])

    # ── Sheet 2: Xulosa ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Xulosa")
    _set_title(ws2, "Xulosa", f"{date_from} — {date_to}")
    summary_rows = [
        ("Jami sotuvlar soni",   summary["total_sales"]),
        ("Jami tushum",          float(summary["total_revenue"])),
        ("Chegirmalar",          float(summary["total_discount"])),
        ("Naqd to'lovlar",       float(summary["cash_total"])),
        ("Karta to'lovlar",      float(summary["card_total"])),
        ("Nasiya",               float(summary["debt_total"])),
        ("Brutto foyda",         float(summary["gross_profit"])),
        ("Foyda margin %",       summary["profit_margin"]),
    ]
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    for r, (label, value) in enumerate(summary_rows, 4):
        ws2.cell(row=r, column=1, value=label).font  = Font(bold=True)
        ws2.cell(row=r, column=2, value=value).alignment = RIGHT

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"

    return _to_buffer(wb)


# ─── Stock Excel ───────────────────────────────────────────────────────────────

def export_stock_excel(stock_qs, summary: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ombor holati"
    _set_title(ws, "Ombor holati hisoboti")

    headers = [
        ("№",           5),  ("Mahsulot",      30), ("Kategoriya",   18),
        ("Birlik",       8),  ("Qoldiq",        12), ("Min qoldiq",   12),
        ("Holat",       12),  ("Tan narxi",     14), ("Sotish narxi", 14),
        ("Tan qiymati", 16),  ("Sotuv qiymati", 16),
    ]
    _set_header_row(ws, 4, headers)

    row = 5
    for idx, stock in enumerate(stock_qs, 1):
        p       = stock.product
        is_low  = stock.quantity <= p.min_stock
        status  = "⚠️ Kam" if is_low else ("❌ Tugagan" if stock.quantity <= 0 else "✅ Yetarli")

        _data_cell(ws, row, 1,  idx,                                   CENTER)
        _data_cell(ws, row, 2,  p.name,                                LEFT)
        _data_cell(ws, row, 3,  p.category.name if p.category else "", LEFT)
        _data_cell(ws, row, 4,  p.unit.short_name if p.unit else "",   CENTER)
        _data_cell(ws, row, 5,  float(stock.quantity),                 RIGHT, '#,##0.00')
        _data_cell(ws, row, 6,  float(p.min_stock),                    RIGHT, '#,##0.00')
        _data_cell(ws, row, 7,  status,                                CENTER)
        _data_cell(ws, row, 8,  float(p.cost_price),                   RIGHT, '#,##0.00')
        _data_cell(ws, row, 9,  float(p.sell_price),                   RIGHT, '#,##0.00')
        _data_cell(ws, row, 10, float(stock.quantity * p.cost_price),  RIGHT, '#,##0.00')
        _data_cell(ws, row, 11, float(stock.quantity * p.sell_price),  RIGHT, '#,##0.00')
        row += 1

    _total_row(ws, row, [
        (1,  "JAMI"),
        (2,  f"{summary['total_items']} ta mahsulot"),
        (10, float(summary["total_cost_value"])),
        (11, float(summary["total_sell_value"])),
    ])

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
    return _to_buffer(wb)


# ─── Purchases Excel ───────────────────────────────────────────────────────────

def export_purchases_excel(purchases_qs, date_from, date_to) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xaridlar"
    _set_title(ws, "Xaridlar hisoboti", f"{date_from} — {date_to}")

    headers = [
        ("№",           5),  ("Kompaniya",    25), ("Filial",       20),
        ("Hisob-faktura",18), ("Sana",         14), ("Jami summa",  14),
        ("To'landi",    14),  ("Qarz",         14), ("Muddat",      14),
        ("Holat",       14),
    ]
    _set_header_row(ws, 4, headers)

    row = 5
    total_sum = Decimal("0")
    total_debt = Decimal("0")

    for idx, purchase in enumerate(purchases_qs, 1):
        _data_cell(ws, row, 1,  idx,                                             CENTER)
        _data_cell(ws, row, 2,  purchase.company.name,                           LEFT)
        _data_cell(ws, row, 3,  purchase.branch.name if purchase.branch else "", LEFT)
        _data_cell(ws, row, 4,  purchase.invoice_no or "—",                      CENTER)
        _data_cell(ws, row, 5,  purchase.created_at.strftime("%d.%m.%Y"),        CENTER)
        _data_cell(ws, row, 6,  float(purchase.total_amount),                    RIGHT, '#,##0.00')
        _data_cell(ws, row, 7,  float(purchase.paid_amount),                     RIGHT, '#,##0.00')
        _data_cell(ws, row, 8,  float(purchase.debt_amount),                     RIGHT, '#,##0.00')
        _data_cell(ws, row, 9,  str(purchase.due_date) if purchase.due_date else "—", CENTER)
        _data_cell(ws, row, 10, purchase.get_status_display(),                   CENTER)
        total_sum  += purchase.total_amount
        total_debt += purchase.debt_amount
        row += 1

    _total_row(ws, row, [
        (1, "JAMI"), (6, float(total_sum)), (8, float(total_debt))
    ])

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
    return _to_buffer(wb)


# ─── Top Products Excel ────────────────────────────────────────────────────────

def export_top_products_excel(data, date_from, date_to) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top mahsulotlar"
    _set_title(ws, "Eng ko'p sotilgan mahsulotlar", f"{date_from} — {date_to}")

    headers = [
        ("№", 5), ("Mahsulot", 35), ("Birlik", 8),
        ("Miqdor", 12), ("Tushum", 16), ("Foyda", 16),
    ]
    _set_header_row(ws, 4, headers)

    row = 5
    for idx, item in enumerate(data, 1):
        _data_cell(ws, row, 1, idx,                                  CENTER)
        _data_cell(ws, row, 2, item["product__name"],                LEFT)
        _data_cell(ws, row, 3, item["product__unit__short_name"] or "", CENTER)
        _data_cell(ws, row, 4, float(item["total_qty"]),             RIGHT, '#,##0.00')
        _data_cell(ws, row, 5, float(item["total_revenue"] or 0),   RIGHT, '#,##0.00')
        _data_cell(ws, row, 6, float(item["total_profit"]  or 0),   RIGHT, '#,##0.00')
        row += 1

    ws.freeze_panes = "A5"
    return _to_buffer(wb)
